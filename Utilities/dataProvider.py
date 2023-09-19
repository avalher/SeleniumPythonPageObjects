import openpyxl


def get_data(sheetName):

    workbook = openpyxl.load_workbook("..//excel//TestData.xlsx")
    # The name of the sheet of the book
    sheet = workbook[sheetName]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    main_list = []

    for row_file in range(2, totalrows + 1):
        datalist = []
        for column_file in range(1, totalcols + 1):
            data = sheet.cell(row=row_file, column=column_file).value
            datalist.insert(column_file, data)
        main_list.insert(row_file, datalist)
    print(main_list)
    return main_list
